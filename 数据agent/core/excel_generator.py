"""
Excel生成模块
将OCR识别的全量数据转换为Excel格式
"""

import re
from typing import Dict, List, Optional
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class ExcelGenerator:
    """Excel文件生成器"""
    
    def __init__(self):
        """初始化Excel生成器"""
        pass
    
    def _clean_text(self, text: str) -> str:
        """清理文本，移除多余空格和特殊字符"""
        text = re.sub(r'[\*\#]+', '', text)
        text = ' '.join(text.split())
        return text.strip()
    
    def _is_valid_item(self, item: Dict, header_keywords: List[str], patient_info_keywords: List[str]) -> bool:
        """检查项目是否有效"""
        project_name = item['项目名称']
        
        # 跳过空项目名称
        if not project_name or project_name in ['', '|', '---']:
            return False
        
        # 跳过表头和无效数据
        if any(kw in project_name for kw in header_keywords):
            return False
        
        # 跳过患者信息
        if any(kw in project_name for kw in patient_info_keywords):
            return False
        
        # 检查是否有结果值或至少有项目名称
        result = item.get('结果', '')
        if not result and not project_name:
            return False
        
        return True
    
    def _parse_html_table(self, html_text: str) -> List[Dict]:
        """
        解析HTML表格格式
        
        Args:
            html_text: 包含HTML表格的文本
            
        Returns:
            List[Dict]: 解析后的数据列表
        """
        data = []
        
        try:
            # 提取所有表格行
            rows = re.findall(r'<tr>(.*?)</tr>', html_text, re.DOTALL)
            
            if len(rows) < 2:  # 至少需要表头和一行数据
                return []
            
            # 解析表头
            header_cells = re.findall(r'<td>(.*?)</td>', rows[0])
            
            # 查找列索引
            project_col = None
            result_col = None
            ref_range_col = None
            unit_col = None
            
            for idx, cell in enumerate(header_cells):
                cell_lower = cell.lower()
                if '项目' in cell_lower or '项目名称' in cell_lower:
                    project_col = idx
                elif '结果' in cell_lower:
                    result_col = idx
                elif '参考' in cell_lower or '区间' in cell_lower:
                    ref_range_col = idx
                elif '单位' in cell_lower:
                    unit_col = idx
            
            # 如果找不到必要的列，使用默认索引
            if project_col is None:
                project_col = 1  # 通常是第2列
            if result_col is None:
                result_col = 3  # 通常是第4列
            if ref_range_col is None:
                ref_range_col = 4  # 通常是第5列
            if unit_col is None:
                unit_col = 5  # 通常是第6列
            
            # 解析数据行（跳过表头）
            for row in rows[1:]:
                cells = re.findall(r'<td>(.*?)</td>', row)
                
                if len(cells) <= project_col:
                    continue
                
                # 提取数据
                item = {
                    '项目名称': self._clean_text(cells[project_col]) if project_col < len(cells) else '',
                    '原始数据': row
                }
                
                # 提取结果值
                if result_col < len(cells):
                    result_value = self._clean_text(cells[result_col])
                    # 检查是否需要从下一行获取箭头标记
                    if result_value in ['↑', '↓', '←', '→'] and result_col + 1 < len(cells):
                        next_value = self._clean_text(cells[result_col + 1])
                        if re.match(r'^\d+\.?\d*$', next_value):
                            result_value = f"{result_value} {next_value}"
                    item['结果'] = result_value
                
                # 提取参考范围
                if ref_range_col < len(cells):
                    ref_range = self._clean_text(cells[ref_range_col])
                    if ref_range:
                        item['参考范围'] = ref_range
                
                # 提取单位
                if unit_col < len(cells):
                    unit = self._clean_text(cells[unit_col])
                    if unit:
                        item['单位'] = unit
                
                # 验证并添加项目
                header_keywords = ['项目名称', '结果', '参考范围', '单位', '---']
                patient_info_keywords = ['姓名', '性别', '年龄', '住院号', '门诊号', '样本号']
                if self._is_valid_item(item, header_keywords, patient_info_keywords):
                    data.append(item)
            
        except Exception as e:
            # HTML解析失败，返回空列表
            pass
        
        return data
    
    def _parse_ocr_text(self, full_text: str) -> List[Dict]:
        """
        解析OCR文本，提取结构化数据（增强版本，支持HTML表格）
        
        Args:
            full_text: OCR识别的全量文本
            
        Returns:
            List[Dict]: 解析后的数据列表
        """
        # 检查是否为HTML表格格式
        if '<table>' in full_text or '<tr>' in full_text:
            html_data = self._parse_html_table(full_text)
            if html_data:
                return html_data
        
        # 定义本地辅助函数
        def clean_text(text: str) -> str:
            """清理文本，移除多余空格和特殊字符"""
            text = re.sub(r'[\*\#]+', '', text)
            text = ' '.join(text.split())
            return text.strip()
        
        def extract_item_from_parts(parts: List[str], start_idx: int) -> Dict:
            """从parts列表中提取一个项目"""
            item = {
                '项目名称': clean_text(parts[start_idx]) if start_idx < len(parts) else '',
                '结果': clean_text(parts[start_idx + 1]) if start_idx + 1 < len(parts) else '',
                '参考范围': clean_text(parts[start_idx + 2]) if start_idx + 2 < len(parts) else '',
                '单位': clean_text(parts[start_idx + 3]) if start_idx + 3 < len(parts) else ''
            }
            return item
        
        def is_valid_item_local(item: Dict) -> bool:
            """检查项目是否有效（本地版本，增强过滤）"""
            project_name = item.get('项目名称', '')
            result = item.get('结果', '')
            
            # 基本检查：空项目名称
            if not project_name or project_name in ['', '|', '---']:
                return False
            
            # 过滤表头关键词
            header_keywords = ['项目名称', '结果', '参考范围', '单位', '---', 'No', 'Test']
            if any(kw in project_name for kw in header_keywords):
                return False
            
            # 过滤患者信息
            patient_info_keywords = ['姓名', '性别', '年龄', '住院号', '门诊号', '样本号']
            if any(kw in project_name for kw in patient_info_keywords):
                return False
            
            # 过滤数字+点号格式（如"21."、"22."）
            if re.match(r'^\d+\.$', project_name):
                return False
            
            # 过滤以数字开头+中文说明的格式（如"6. 对于..."）
            if re.match(r'^\d+\.', project_name):
                return False
            
            # 过滤以数字开头+中文说明的格式（如"6. 对于..."）
            if re.match(r'^\d+\.\s*[\u4e00-\u9fff]', project_name):
                return False
            
            # 过滤纯数字项目名称（除非是常见的医学指标编号）
            if project_name.isdigit() and len(project_name) > 2:
                return False
            
            # 检查结果值是否为中文说明文字
            if result and re.search(r'[\u4e00-\u9fff]{5,}', result):
                return False
            
            # 检查结果值是否为说明性文字
            if result and any(phrase in result for phrase in ['请确保', '格式正确', '样本类型', '英文数字', '特殊字符', '转义字符']):
                return False
            
            # 检查是否有结果值或至少有项目名称
            if not result and not project_name:
                return False
            
            # 项目名称应该是中文或医学术语
            if project_name and re.match(r'^[a-zA-Z\s]{10,}$', project_name):
                return False
            
            return True
            
            # 新增：检查结果值是否为说明性文字
            if result and any(phrase in result for phrase in ['请确保', '格式正确', '样本类型', '英文数字', '特殊字符', '转义字符']):
                return False
            
            # 检查是否有结果值或至少有项目名称
            if not result and not project_name:
                return False
            
            # 新增：项目名称应该是中文或医学术语
            # 排除纯英文说明文字
            if project_name and re.match(r'^[a-zA-Z\s]{10,}$', project_name):
                return False
            
            return True
        
        lines = full_text.strip().split('\n')
        data = []
        
        # 增强的表头关键词集合（中英文）
        header_keywords = [
            '项目名称', '结果', '参考范围', '单位', 'Test', 'Result', 
            'Reference', 'Unit', 'Normal', 'Range', 'Value', '---',
            '检验项目', '检测项目', '测定值', '正常值', '异常'
        ]
        
        # 患者信息关键词（需要跳过）
        patient_info_keywords = [
            '姓名', '性别', '年龄', '住院号', '门诊号', '样本号',
            '标本号', '送检日期', '检验日期', '科室', '床号',
            'Name', 'Sex', 'Age', 'ID', 'Sample', 'Specimen',
            'Date', 'Department'
        ]
        
        # 格式标记（需要跳过）
        format_markers = ['###', '##', '****', '***', '**', '==', '---', '...']
        
        def is_header_line(line: str) -> bool:
            """判断是否为表头行"""
            # 检查是否包含多个表头关键词
            keyword_count = sum(1 for kw in header_keywords if kw in line)
            if keyword_count >= 2:
                return True
            
            # 检查是否以格式标记开头
            for marker in format_markers:
                if line.startswith(marker):
                    return True
            
            # 检查是否只包含分隔符
            if re.match(r'^[|\s\-+=]+$', line):
                return True
            
            return False
        
        def is_patient_info(line: str) -> bool:
            """判断是否为患者信息行"""
            for keyword in patient_info_keywords:
                if keyword in line:
                    return True
            return False
        
        def extract_item_from_parts(parts: List[str], start_idx: int) -> Dict:
            """从parts列表中提取一个项目"""
            item = {
                '项目名称': self._clean_text(parts[start_idx]) if start_idx < len(parts) else '',
                '结果': self._clean_text(parts[start_idx + 1]) if start_idx + 1 < len(parts) else '',
                '参考范围': self._clean_text(parts[start_idx + 2]) if start_idx + 2 < len(parts) else '',
                '单位': self._clean_text(parts[start_idx + 3]) if start_idx + 3 < len(parts) else ''
            }
            return item
        
        def is_valid_item(item: Dict) -> bool:
            """检查项目是否有效（增强版本）"""
            project_name = item.get('项目名称', '')
            result = item.get('结果', '')
            
            # 基本检查：空项目名称
            if not project_name or project_name in ['', '|', '---']:
                return False
            
            # 过滤表头关键词
            header_keywords = ['项目名称', '结果', '参考范围', '单位', '---', 'No', 'Test']
            if any(kw in project_name for kw in header_keywords):
                return False
            
            # 过滤患者信息
            patient_info_keywords = ['姓名', '性别', '年龄', '住院号', '门诊号', '样本号']
            if any(kw in project_name for kw in patient_info_keywords):
                return False
            
            # 过滤数字+点号格式（如"21."、"22."）
            if re.match(r'^\d+\.$', project_name):
                return False
            
            # 过滤以数字开头的格式（如"6. 对于..."、"222. 对于..."）
            if re.match(r'^\d+\.', project_name):
                return False
            
            # 过滤纯数字项目名称
            if project_name.isdigit() and len(project_name) > 2:
                return False
            
            # 检查结果值是否为中文说明文字
            if result and re.search(r'[\u4e00-\u9fff]{5,}', result):
                return False
            
            # 检查结果值是否为说明性文字
            if result and any(phrase in result for phrase in ['请确保', '格式正确', '样本类型', '英文数字', '特殊字符', '转义字符', '多行数据', '制表符', '空格分隔']):
                return False
            
            # 检查项目名称是否为纯英文说明
            if project_name and re.match(r'^[a-zA-Z\s]{10,}$', project_name):
                return False
            
            # 检查是否有结果值或至少有项目名称
            if not result and not project_name:
                return False
            
            return True
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # 跳过表头行
            if is_header_line(line):
                continue
            
            # 跳过患者信息行
            if is_patient_info(line):
                continue
            
            # 优先使用竖线分隔符（表格格式）
            if '|' in line:
                parts = [p.strip() for p in line.split('|')]
                # 过滤掉空的部分
                parts = [p for p in parts if p]
                
                if len(parts) >= 2:
                    items_from_line = []
                    
                    # 尝试解析项目
                    if len(parts) >= 4:
                        # 标准格式：项目名称|结果|参考范围|单位
                        item = extract_item_from_parts(parts, 0)
                        if is_valid_item(item):
                            item['原始数据'] = line
                            items_from_line.append(item)
                        
                        # 检查是否有更多项目（多行合并的情况）
                        # 从索引4开始，循环处理剩余的项目
                        current_idx = 4
                        while current_idx < len(parts):
                            remaining_parts = len(parts) - current_idx
                            
                            if remaining_parts < 2:
                                # 剩余部分不足，无法构成有效项目
                                break
                            
                            # 根据剩余部分数量判断项目格式
                            next_item = None
                            
                            if remaining_parts >= 4:
                                # 检查是否是4列项目还是3列项目+下一项目的开始
                                # 通过检查第4列（索引current_idx+3）是否是有效项目名称来判断
                                potential_unit = clean_text(parts[current_idx + 3])
                                potential_next_name = clean_text(parts[current_idx + 4]) if remaining_parts >= 5 else ''
                                
                                # 更严格的单位识别
                                # 单位通常包含：数字、科学计数法、%、/、等特殊字符
                                is_likely_unit = bool(
                                    re.match(r'^[%a-zA-Z/]+$', potential_unit) and len(potential_unit) <= 8 or  # 简单单位，短
                                    re.match(r'^×10\^?\d*/[a-zA-Z]+$', potential_unit) or  # 科学计数法单位
                                    re.match(r'^[a-zA-Z]{1,3}/[a-zA-Z]{1,3}$', potential_unit) or  # 斜杠分隔的单位
                                    re.match(r'^\d+[a-zA-Z]+$', potential_unit)  # 数字+字母（如10^9）
                                )
                                
                                # 如果第4列看起来像单位，且第5列存在且不像项目名称，则认为是4列项目
                                # 否则认为是3列项目
                                if is_likely_unit and remaining_parts == 4:
                                    # 确定是4列项目
                                    next_item = extract_item_from_parts(parts, current_idx)
                                    current_idx += 4
                                else:
                                    # 3列项目，第4列是下一个项目的开始
                                    third_column = clean_text(parts[current_idx + 2])
                                    is_reference_range = bool(
                                        re.match(r'^\d+\.?\d*[-~]\d+\.?\d*$', third_column) or
                                        re.match(r'^[<>≤≥]\d+\.?\d*$', third_column) or
                                        re.match(r'^[<>≤≥]=?\d+\.?\d*$', third_column)
                                    )
                                    
                                    if is_reference_range:
                                        next_item = {
                                            '项目名称': clean_text(parts[current_idx]),
                                            '结果': clean_text(parts[current_idx + 1]),
                                            '参考范围': third_column,
                                            '原始数据': line
                                        }
                                    else:
                                        next_item = {
                                            '项目名称': clean_text(parts[current_idx]),
                                            '结果': clean_text(parts[current_idx + 1]),
                                            '单位': third_column,
                                            '原始数据': line
                                        }
                                    current_idx += 3
                            elif remaining_parts == 3:
                                # 3列项目，需要判断第三列是参考范围还是单位
                                third_column = clean_text(parts[current_idx + 2])
                                is_reference_range = bool(
                                    re.match(r'^\d+\.?\d*[-~]\d+\.?\d*$', third_column) or
                                    re.match(r'^[<>≤≥]\d+\.?\d*$', third_column) or
                                    re.match(r'^[<>≤≥]=?\d+\.?\d*$', third_column)
                                )
                                
                                if is_reference_range:
                                    next_item = {
                                        '项目名称': clean_text(parts[current_idx]),
                                        '结果': clean_text(parts[current_idx + 1]),
                                        '参考范围': third_column,
                                        '原始数据': line
                                    }
                                else:
                                    next_item = {
                                        '项目名称': clean_text(parts[current_idx]),
                                        '结果': clean_text(parts[current_idx + 1]),
                                        '单位': third_column,
                                        '原始数据': line
                                    }
                                current_idx += 3
                            else:
                                # 2列项目：项目名称|结果
                                next_item = {
                                    '项目名称': clean_text(parts[current_idx]),
                                    '结果': clean_text(parts[current_idx + 1]),
                                    '原始数据': line
                                }
                                current_idx += 2
                            
                            # 验证并添加项目
                            if next_item and is_valid_item(next_item):
                                items_from_line.append(next_item)
                    elif len(parts) == 2:
                        # 简化格式：项目名称|结果
                        item = {
                            '项目名称': clean_text(parts[0]),
                            '结果': clean_text(parts[1]),
                            '原始数据': line
                        }
                        if is_valid_item(item):
                            items_from_line.append(item)
                    elif len(parts) == 3:
                        # 三列格式：需要智能判断是"项目名称|结果|参考范围"还是"项目名称|结果|单位"
                        third_column = clean_text(parts[2])
                        
                        # 判断第三列是否为参考范围
                        # 参考范围格式：
                        # 1. 数字范围：如"0.4-8.0"、"3.5-9.5"
                        # 2. 小于/大于：如"<10"、">5.0"、"≤50"
                        # 3. 其他格式：如"阴性"、"阳性"
                        is_reference_range = bool(
                            re.match(r'^\d+\.?\d*[-~]\d+\.?\d*$', third_column) or  # 数字范围
                            re.match(r'^[<>≤≥]\d+\.?\d*$', third_column) or       # 小于/大于
                            re.match(r'^[<>≤≥]=?\d+\.?\d*$', third_column)        # 小于等于/大于等于
                        )
                        
                        if is_reference_range:
                            # 格式：项目名称|结果|参考范围
                            item = {
                                '项目名称': clean_text(parts[0]),
                                '结果': clean_text(parts[1]),
                                '参考范围': third_column,
                                '原始数据': line
                            }
                        else:
                            # 格式：项目名称|结果|单位
                            item = {
                                '项目名称': clean_text(parts[0]),
                                '结果': clean_text(parts[1]),
                                '单位': third_column,
                                '原始数据': line
                            }
                        
                        if is_valid_item(item):
                            items_from_line.append(item)
                    
                    # 将有效的项目添加到数据列表
                    for item in items_from_line:
                        data.append(item)
                continue
            
            # 如果没有竖线，尝试冒号分隔符
            if ':' in line or '：' in line:
                parts = re.split(r'[:：]', line, 1)
                if len(parts) == 2:
                    item = {
                        '项目名称': clean_text(parts[0]),
                        '结果': clean_text(parts[1]),
                        '原始数据': line
                    }
                    if is_valid_item(item):
                        data.append(item)
                continue
            
            # 尝试制表符分隔
            if '\t' in line:
                parts = [p.strip() for p in line.split('\t')]
                parts = [p for p in parts if p]
                
                if len(parts) >= 2:
                    item = {
                        '项目名称': clean_text(parts[0]),
                        '结果': clean_text(parts[1]),
                        '原始数据': line
                    }
                    if len(parts) >= 3:
                        item['参考范围'] = clean_text(parts[2])
                    if len(parts) >= 4:
                        item['单位'] = clean_text(parts[3])
                    
                    if is_valid_item(item):
                        data.append(item)
                continue
            
# 最后尝试空格分隔（包括特殊格式：序号 项目名称 医院标识 结果 参考范围 单位）
            parts = re.split(r'\s+', line)
            
            if len(parts) >= 2:
                # 检查是否是特殊格式：序号 项目名称 医院标识 结果 参考范围 单位
                # 特征：第二个部分包含【】（医院标识）
                is_special_format = False
                if len(parts) >= 4:
                    first_part = parts[0]
                    second_part = parts[1] if len(parts) > 1 else ''
                    
                    # 检查第一个部分是否为数字（序号）
                    is_number = bool(re.match(r'^\d+$', first_part))
                    # 检查第二个部分是否包含医院标识（【】）
                    has_hospital_marker = '【' in second_part and '】' in second_part
                    
                    if is_number and has_hospital_marker:
                        is_special_format = True
                
                if is_special_format:
                    # 特殊格式解析：序号 医院标识 项目名称 [箭头] 结果 参考范围 单位 检测方法
                    item = {
                        '项目名称': clean_text(parts[2]),  # 项目名称（在索引2）
                        '原始数据': line
                    }
                    
                    # 查找结果值（从第4个部分开始，跳过序号、项目名称、医院标识）
                    result_value = None
                    result_index = 3
                    arrow_marker = None  # 存储箭头标记
                    
                    # 检查第4个部分是否为箭头标记
                    if result_index < len(parts) and parts[result_index] in ['↑', '↓', '←', '→']:
                        arrow_marker = parts[result_index]
                        result_index += 1
                    
                    # 查找数字结果值（跳过可能的医院标识）
                    while result_index < len(parts):
                        part = parts[result_index]
                        # 检查是否为数字（跳过医院标识如【深圳HR】）
                        if re.match(r'^\d+\.?\d*$', part):
                            result_value = part
                            break
                        result_index += 1
                    
                    if result_value:
                        # 组合箭头标记和结果值
                        if arrow_marker:
                            item['结果'] = f"{arrow_marker} {result_value}"
                        else:
                            item['结果'] = result_value
                        
                        # 查找参考范围
                        ref_range = None
                        ref_index = result_index + 1
                        while ref_index < len(parts):
                            part = parts[ref_index]
                            if re.match(r'^\d+\.?\d*[-~]\d+\.?\d*$', part):
                                ref_range = part
                                break
                            ref_index += 1
                        
                        if ref_range:
                            item['参考范围'] = ref_range
                            
                            # 查找单位
                            unit = None
                            unit_index = ref_index + 1
                            while unit_index < len(parts):
                                part = parts[unit_index]
                                # 单位通常包含字母、/、数字等
                                if re.match(r'^[a-zA-Z/\d^%]+$', part):
                                    unit = part
                                    break
                                # 检查是否是检测方法（通常以"法"结尾）
                                if part.endswith('法') or part.endswith('法)'):
                                    break
                                unit_index += 1
                            
                            if unit:
                                item['单位'] = unit
                        else:
                            # 没有找到参考范围，尝试查找单位
                            unit = None
                            unit_index = result_index + 1
                            while unit_index < len(parts):
                                part = parts[unit_index]
                                if re.match(r'^[a-zA-Z/\d^%]+$', part):
                                    unit = part
                                    break
                                if part.endswith('法') or part.endswith('法)'):
                                    break
                                unit_index += 1
                            
                            if unit:
                                item['单位'] = unit
                else:
                    # 普通空格分隔格式
                    item = {
                        '项目名称': parts[0],
                        '原始数据': line
                    }
                    
                    # 尝试提取结果值
                    if len(parts) >= 2:
                        item['结果'] = parts[1]
                    
                    # 尝试提取参考范围
                    if len(parts) >= 3:
                        item['参考范围'] = parts[2]
                    
                    # 尝试提取单位
                    if len(parts) >= 4:
                        item['单位'] = parts[3]
                
                if is_valid_item(item):
                    data.append(item)
        
        # 如果没有解析到结构化数据，返回原始文本
        if not data:
            return [{'原始数据': full_text}]
        
        return data
    
    def generate_excel(
        self, 
        full_text: str, 
        output_path: Optional[str] = None
    ) -> str:
        """
        生成Excel文件
        
        Args:
            full_text: OCR识别的全量文本
            output_path: 输出文件路径，如果为None则使用默认路径
            
        Returns:
            str: 生成的Excel文件路径
        """
        # 设置默认输出路径（使用时间戳避免文件冲突）
        if output_path is None:
            output_dir = Path(__file__).parent.parent / "output"
            output_dir.mkdir(exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = output_dir / f"medical_data_{timestamp}.xlsx"
        
        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "医学数据"
        
        # 设置样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 解析数据
        data = self._parse_ocr_text(full_text)
        
        # 写入表头
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 写入数据
        for row_idx, item in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=item.get(header, ''))
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存文件
        wb.save(output_path)
        
        return str(output_path)
    
    def generate_full_data_json(self, full_text: str) -> Dict:
        """
        生成全量数据的JSON格式
        
        Args:
            full_text: OCR识别的全量文本
            
        Returns:
            Dict: 包含全量数据的字典
        """
        data = self._parse_ocr_text(full_text)
        
        return {
            'raw_text': full_text,
            'structured_data': data,
            'item_count': len(data)
        }
